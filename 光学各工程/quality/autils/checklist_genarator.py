# 特别鸣谢：chatgpt3.5

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill,
    Border,
    Side,
    Alignment,
    Protection,
    Font
)
from openpyxl.styles import NamedStyle
from openpyxl.styles.builtins import styles
import datetime
import random

def random_str(slen=10):
    seed = "1234567890abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ"
    sa = []
    for i in range(slen):
      sa.append(random.choice(seed))
    return ''.join(sa)
time_str = "{}_{}".format(datetime.datetime.now().strftime('_%Y_%m_%d_%H_%M'),random_str(4))

#######################################################################################

wb = Workbook()
# TODO：0.预填项

tbd_ver = 'norm404079300 hard 404079300'
tbd_note = '测试自动生成checklist'
tbd_name = '潘琪'
tbd_path_save_xlsx = './算法网络checklist_'+tbd_ver+time_str+'.xlsx'

################################
ver = '版本号信息'
note = '更新说明'
time = '耗时变化'
memory = '内存堆栈变化'
leak = '内存泄露测试'
phone = '整机端测试建议'
test = '是否有离线跑库数据'

#######################################################################################

# TODO：1.建立工作表。后续新增工作表在此处新建。
ws0 = wb.create_sheet("算法网络checklist", 0)
ws1 = wb.create_sheet("离线跑库数据", 1)

ws0.sheet_properties.tabColor = "005CAF"#琉璃
ws1.sheet_properties.tabColor = "1B813E"#常磐

print('生成工作表如下： ')
print(wb.sheetnames)

style_calculation = NamedStyle(name='calculation_style', builtinId=6)
style_check_cell = NamedStyle(name='check_cell', builtinId=23)

# wb.add_named_style(style_calculation)
# wb.add_named_style(style_check_cell)
# ws0['A1'].style=style_calculation
# ws0['B1'].style=style_calculation
# ws0['C1'].style=style_calculation
# ws0['D1'].style=style_calculation
# ws0['A14'].style=style_check_cell
# ws0['C14'].style=style_check_cell
# ws0['C15'].style=style_check_cell
ws0['A1'].style='Check Cell'
ws0['B1'].style='Check Cell'
ws0['C1'].style='Check Cell'
ws0['D1'].style='Check Cell'
ws0['A14'].style='Calculation'
ws0['C14'].style='Calculation'
ws0['C15'].style='Calculation'

#######################################################################################

# TODO：2.合并单元格。后续新增其他检查项，以及修改当前检查项在此处修改。
ws0.merge_cells('A2:A5')    #输入信息确认
ws0.merge_cells('A6:A13')   #异物网络
ws0.merge_cells('A14:A21')  #防伪网络
ws0.merge_cells('A22:A29')  #残留网络
ws0.merge_cells('A30:A37')  #误触网络
ws0.merge_cells('A38:A45')  #提点网络

ws0.merge_cells('D2:D5')    #name
ws0.merge_cells('D6:D13')   #name
ws0.merge_cells('D14:D21')  #name
ws0.merge_cells('D22:D29')  #name
ws0.merge_cells('D30:D37')  #name
ws0.merge_cells('D38:D45')  #name


ws0['A2'] = '输入信息确认'
ws0['A6'] = '异物网络'
ws0['A14'] = '防伪网络'
ws0['A22'] = '残留网络'
ws0['A30'] = '增强网络'
ws0['A38'] = '提点网络'

#######################################################################################

# TODO：3.整体结构。后续修改基本骨架在此处修改。

ws0['B1'] = '版本发布检查项'
ws0['B2'] = '常规图库(常温、低温、强光、洗手、部分按压、脏污划痕、贴膜)是否齐全'
ws0['B3'] = '网络图库(异物防伪等)是否齐全'
ws0['B4'] = '所报问题(图像、log)是否齐全'
ws0['B5'] = '是否可支撑版本优化'
ws0['C1'] = '结果'
ws0['D1'] = '填写人员'

## 异物
ws0['B6'] = ver
ws0['B7'] = note
ws0['B8'] = time
ws0['B9'] = memory
ws0['B10'] = leak
ws0['B11'] = phone
ws0['B12'] = test

## 防伪
ws0['B14'] = ver
ws0['B15'] = note
ws0['B16'] = time
ws0['B17'] = memory
ws0['B18'] = leak
ws0['B19'] = phone
ws0['B20'] = test

## 残留
ws0['B22'] = ver
ws0['B23'] = note
ws0['B24'] = time
ws0['B25'] = memory
ws0['B26'] = leak
ws0['B27'] = phone
ws0['B28'] = test

## 误触
ws0['B30'] = ver
ws0['B31'] = note
ws0['B32'] = time
ws0['B33'] = memory
ws0['B34'] = leak
ws0['B35'] = phone
ws0['B36'] = test

## 提点
ws0['B38'] = ver
ws0['B39'] = note
ws0['B40'] = time
ws0['B41'] = memory
ws0['B42'] = leak
ws0['B43'] = phone
ws0['B44'] = test

#######################################################################################

# TODO：4.表1预填项生成。
ws0['C2'] = '齐全'
ws0['C3'] = '齐全'
ws0['C4'] = '齐全'
ws0['C5'] = '是'

## 防伪为例
# ver = '版本号信息'
# note = '更新说明'
# time = '耗时变化'
# memory = '内存堆栈变化'
# leak = '内存泄露测试'
# phone = '整机端测试建议'
# test = '是否有离线跑库数据'
#不做修改
ws0['C16'] = '无'
ws0['C17'] = '无'
ws0['C18'] = '通过'
ws0['C19'] = '无'
ws0['C20'] = '是'

# 版本号、更新内容、更新者
ws0['C14'] = tbd_ver
ws0['C15'] = tbd_note
ws0['D2'] = tbd_name
ws0['D14'] = tbd_name

#######################################################################################

# TODO：5.表2跑库数据由跑库脚本生成表格流。

#######################################################################################

# TODO: 6 格式

# Set border for all cells
border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))

for row in ws0.iter_rows(min_row=1, max_row=45, min_col=1, max_col=4):
    for cell in row:
        cell.border = border

# Set alignment for all cells
for row in ws0.iter_rows(min_row=1, max_row=45, min_col=1, max_col=4):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Set column width and row height for all cells
for col in ws0.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.6
    ws0.column_dimensions[column].width = adjusted_width

for row in ws0.rows:
    ws0.row_dimensions[row[0].row].height = 20

#Apply the Calculation and CheckCell style to the cell


wb.save(tbd_path_save_xlsx)

import torch
import numpy as np
import torchvision
from torchvision import transforms as transforms

from torch.autograd import Variable

import tqdm
import string
import datetime

import itertools

# import matplotlib.pyplot as plt
import os
from shutil import copyfile
from shutil import move
from math import exp
import torch.nn as nn
import torchvision.transforms.functional as TF
from collections import OrderedDict
import torch.nn as nn

# import pandas as pd
import time as ti
import argparse

# 显示所有列
# pd.set_option('display.max_columns',None)
# 显示所有行
# pd.set_option('display.max_rows',None)
import os

# os.environ["CUDA_VISIBLE_DEVICES"] = "2 "
from torch.autograd import Variable

# from MobileNet import MNV30811
from tqdm import tqdm
import datetime
from pathlib import Path
import itertools
import matplotlib.pyplot as plt
import os
from shutil import copyfile

args = None
from shutil import move
from pathlib import Path
from math import exp
import random
import re
import xlsxwriter
import torch
import numpy as np
import torchvision
from torchvision import transforms as transforms
from mobilenetv3_alignmix import (
    MNV3_large2,
    MNV3_large2_mixup,
    MNV3_large2_drop,
    MNV3_large2_crelu,
)  # ,ImageClassifierHead

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill,
    Border,
    Side,
    Alignment,
    Protection,
    Font,
    NamedStyle,
    NumberFormatDescriptor,
)

from openpyxl.drawing.image import Image as pyxlImage

# from openpyxl.drawing import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles.builtins import styles
import datetime
import itertools

# import matplotlib.pyplot as plt
import os
import shutil
from shutil import copyfile
from shutil import move
from math import exp
import torch.nn as nn
import cv2
import numpy as np
from PIL import Image

global flag_name
import random

alphabet = string.ascii_uppercase

border_thin = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
border_thick = Border(
    left=Side(style="thick"),
    right=Side(style="thick"),
    top=Side(style="thick"),
    bottom=Side(style="thick"),
)


def set_auto_hw(ws, r0, r1):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
        # adjusted_width = (max_length + 2) * 1.4

        # adjusted_width = (max_length + 2) * r0
        ws.column_dimensions[column].width = r0

    for row in ws.rows:
        ws.row_dimensions[row[0].row].height = r1
        # ws.row_dimensions[row[0].row].height = 16


def set_inter_border(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border_thin


def set_inter_align(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )


def set_inter_percent(ws, cell_range):
    for row in ws[cell_range]:
        for cell in row:
            cell.number_format = "0.00%"


def set_outer_border(
    ws, cell_range, isthin
):  # 参数为左上角坐标和右下角坐标，形如'D3','A5'等。ws是worksheet对象。
    start_loc = cell_range.split(":")[0]
    end_loc = cell_range.split(":")[1]
    x_start = start_loc[0]
    y_start = start_loc[1 : len(start_loc)]  # 切片获取坐标的数字部分
    x_end = end_loc[0]
    y_end = end_loc[1 : len(end_loc)]
    len_y = int(y_end) - int(y_start) + 1
    alphabet = string.ascii_uppercase  # 导入字母表
    len_x = alphabet.index(x_end) - alphabet.index(x_start) + 1

    isthin = "thin" if 1 else "thick"
    # 左上
    temp = start_loc
    ws[temp].border = Border(left=Side(style=isthin), top=Side(style=isthin))
    # 右下
    temp = end_loc
    ws[temp].border = Border(right=Side(style=isthin), bottom=Side(style=isthin))
    # 右上
    temp = x_end + y_start
    ws[temp].border = Border(right=Side(style=isthin), top=Side(style=isthin))
    # 左下
    temp = x_start + y_end
    ws[temp].border = Border(left=Side(style=isthin), bottom=Side(style=isthin))
    # 上
    for i in range(0, len_x - 2):
        temp = alphabet[alphabet.index(x_start) + 1 + i] + y_start
        ws[temp].border = Border(top=Side(style=isthin))
    # 下
    for i in range(0, len_x - 2):
        temp = alphabet[alphabet.index(x_start) + 1 + i] + y_end
        ws[temp].border = Border(bottom=Side(style=isthin))
    # 左
    for i in range(0, len_y - 2):
        temp = x_start + str(int(y_start) + 1 + i)
        ws[temp].border = Border(left=Side(style=isthin))
    # 右
    for i in range(0, len_y - 2):
        temp = x_end + str(int(y_start) + 1 + i)
        ws[temp].border = Border(right=Side(style=isthin))
    return 0


# def set_outer_border_loc(ws,min_row,max_row,min_col,max_col,isthin):         #参数为左上角坐标和右下角坐标，形如'D3','A5'等。ws是worksheet对象。

#     x_start = min_col
#     y_start = min_row  #切片获取坐标的数字部分
#     x_end = max_col
#     y_end = max_row
#     len_y = int(y_end) - int(y_start) + 1

#     len_x = int(x_end) - int(x_start) + 1

#     isthin = 'thin' if 1 else 'thick'
#     # 左上
#     temp = start_loc
#     ws[temp].border = Border(left=Side(style=isthin),top=Side(style=isthin))
#     # 右下
#     temp = end_loc
#     ws[temp].border = Border(right=Side(style=isthin),bottom=Side(style=isthin))
#     # 右上
#     temp = x_end + y_start
#     ws[temp].border = Border(right=Side(style=isthin),top=Side(style=isthin))
#     # 左下
#     temp = x_start + y_end
#     ws[temp].border = Border(left=Side(style=isthin),bottom=Side(style=isthin))
#     # 上
#     for i in range(0,len_x-2):
#         temp = alphabet[alphabet.index(x_start)+1+i] + y_start
#         ws[temp].border = Border(top=Side(style=isthin))
#     # 下
#     for i in range(0,len_x-2):
#         temp = alphabet[alphabet.index(x_start)+1+i] + y_end
#         ws[temp].border = Border(bottom=Side(style=isthin))
#     # 左
#     for i in range(0,len_y-2):
#         temp = x_start + str(int(y_start) + 1 + i)
#         ws[temp].border = Border(left=Side(style=isthin))
#     # 右
#     for i in range(0,len_y-2):
#         temp = x_end + str(int(y_start) + 1 + i)
#         ws[temp].border = Border(right=Side(style=isthin))
#     return 0


# 绘制混淆矩阵
def random_str(slen=10):
    seed = "1234567890abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ"
    sa = []
    for i in range(slen):
        sa.append(random.choice(seed))
    return "".join(sa)


time_str = "{}_{}".format(datetime.datetime.now().strftime("_%m%d"), random_str(4))

#######################################################################################

wb = Workbook()
ws0 = wb.create_sheet("算法网络checklist", 0)
ws1 = wb.create_sheet("离线跑库数据", 1)
ws_wuka = wb.create_sheet("误卡图像", 2)
ws_louka = wb.create_sheet("漏卡图像", 3)

ws_sheet = wb["Sheet"]
wb.remove_sheet(ws_sheet)

ws0.sheet_properties.tabColor = "005CAF"  # 琉璃
ws1.sheet_properties.tabColor = "1B813E"  # 常磐
ws_wuka.sheet_properties.tabColor = "F8C3CD"  # 退红
ws_louka.sheet_properties.tabColor = "F8C3CD"  # 退红

print("生成工作表如下： ")
print(wb.sheetnames)


def write_ws0(ws):
    # TODO：0.预填项

    ################################
    ver = "版本号信息"
    note = "更新说明"
    time = "耗时变化"
    memory = "内存堆栈变化"
    leak = "内存泄露测试"
    phone = "整机端测试建议"
    test = "是否有离线跑库数据"

    #######################################################################################

    # TODO：1.建立工作表。后续新增工作表在此处新建。

    style_calculation = NamedStyle(name="calculation_style", builtinId=6)
    style_check_cell = NamedStyle(name="check_cell", builtinId=23)

    # wb.add_named_style(style_calculation)
    # wb.add_named_style(style_check_cell)
    # ws0['A1'].style=style_calculation
    # ws0['B1'].style=style_calculation
    # ws0['C1'].style=style_calculation
    # ws0['D1'].style=style_calculation
    # ws0['A14'].style=style_check_cell
    # ws0['C14'].style=style_check_cell
    # ws0['C15'].style=style_check_cell
    ws0["A1"].style = "Check Cell"
    ws0["B1"].style = "Check Cell"
    ws0["C1"].style = "Check Cell"
    ws0["D1"].style = "Check Cell"
    ws0["A14"].style = "Calculation"
    ws0["C14"].style = "Calculation"
    ws0["C15"].style = "Calculation"

    #######################################################################################

    # TODO：2.合并单元格。后续新增其他检查项，以及修改当前检查项在此处修改。
    ws0.merge_cells("A2:A5")  # 输入信息确认
    ws0.merge_cells("A6:A13")  # 异物网络
    ws0.merge_cells("A14:A21")  # 防伪网络
    ws0.merge_cells("A22:A29")  # 残留网络
    ws0.merge_cells("A30:A37")  # 误触网络
    ws0.merge_cells("A38:A45")  # 提点网络

    ws0.merge_cells("D2:D5")  # name
    ws0.merge_cells("D6:D13")  # name
    ws0.merge_cells("D14:D21")  # name
    ws0.merge_cells("D22:D29")  # name
    ws0.merge_cells("D30:D37")  # name
    ws0.merge_cells("D38:D45")  # name

    ws0["A2"] = "输入信息确认"
    ws0["A6"] = "异物网络"
    ws0["A14"] = "防伪网络"
    ws0["A22"] = "残留网络"
    ws0["A30"] = "增强网络"
    ws0["A38"] = "提点网络"

    #######################################################################################

    # TODO：3.整体结构。后续修改基本骨架在此处修改。

    ws0["B1"] = "版本发布检查项"
    ws0["B2"] = "常规图库(常温、低温、强光、洗手、部分按压、脏污划痕、贴膜)是否齐全"
    ws0["B3"] = "网络图库(异物防伪等)是否齐全"
    ws0["B4"] = "所报问题(图像、log)是否齐全"
    ws0["B5"] = "是否可支撑版本优化"
    ws0["C1"] = "结果"
    ws0["D1"] = "填写人员"

    ## 异物
    ws0["B6"] = ver
    ws0["B7"] = note
    ws0["B8"] = time
    ws0["B9"] = memory
    ws0["B10"] = leak
    ws0["B11"] = phone
    ws0["B12"] = test

    ## 防伪
    ws0["B14"] = ver
    ws0["B15"] = note
    ws0["B16"] = time
    ws0["B17"] = memory
    ws0["B18"] = leak
    ws0["B19"] = phone
    ws0["B20"] = test

    ## 残留
    ws0["B22"] = ver
    ws0["B23"] = note
    ws0["B24"] = time
    ws0["B25"] = memory
    ws0["B26"] = leak
    ws0["B27"] = phone
    ws0["B28"] = test

    ## 误触
    ws0["B30"] = ver
    ws0["B31"] = note
    ws0["B32"] = time
    ws0["B33"] = memory
    ws0["B34"] = leak
    ws0["B35"] = phone
    ws0["B36"] = test

    ## 提点
    ws0["B38"] = ver
    ws0["B39"] = note
    ws0["B40"] = time
    ws0["B41"] = memory
    ws0["B42"] = leak
    ws0["B43"] = phone
    ws0["B44"] = test

    #######################################################################################

    # TODO：4.表1预填项生成。
    ws0["C2"] = "齐全"
    ws0["C3"] = "齐全"
    ws0["C4"] = "齐全"
    ws0["C5"] = "是"

    ## 防伪为例
    # ver = '版本号信息'
    # note = '更新说明'
    # time = '耗时变化'
    # memory = '内存堆栈变化'
    # leak = '内存泄露测试'
    # phone = '整机端测试建议'
    # test = '是否有离线跑库数据'
    # 不做修改
    ws0["C16"] = "无"
    ws0["C17"] = "无"
    ws0["C18"] = "通过"
    ws0["C19"] = "无"
    ws0["C20"] = "是"

    # 版本号、更新内容、更新者
    ws0["C14"] = tbd_ver
    ws0["C15"] = tbd_note
    ws0["D2"] = tbd_name
    ws0["D14"] = tbd_name

    #######################################################################################

    # TODO：5.表2跑库数据由跑库脚本生成表格流。

    #######################################################################################

    # TODO: 6 格式

    # Set border for all cells

    for row in ws0.iter_rows(min_row=1, max_row=45, min_col=1, max_col=4):
        for cell in row:
            cell.border = border_thin

    # Set alignment for all cells
    for row in ws0.iter_rows(min_row=1, max_row=45, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set column width and row height for all cells

    for col in ws0.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.6
        ws0.column_dimensions[column].width = adjusted_width

    for row in ws0.rows:
        ws0.row_dimensions[row[0].row].height = 20

    # Apply the Calculation and CheckCell style to the cell


def parse_args():
    parser = argparse.ArgumentParser(description="ASP Train Network")
    parser.add_argument(
        "-lm", "--listmodelpath", nargs="+", default="", help="list of train dir"
    )
    parser.add_argument(
        "-j", "--jilianmodelpath", nargs="+", default="", help="list of train dir"
    )
    parser.add_argument(
        "-lv", "--listvaliddir", nargs="+", default="", help="list of valid dir"
    )
    # parser.add_argument(
    #     "-j",
    #     "--jilianmodelpath",
    #     default="./STANDALONE/model_STANDALONE_2023_07_10_03_07/ckpt_208.pth",
    #     help="train dir",
    # )
    parser.add_argument("-m", "--modelpath", default="", help="train dir")
    parser.add_argument(
        "-x",
        "--xlsxpath",
        default="/home/panq/tool/aosp/vendor/confusion_xlsx/算法网络checklist_开学习_",
        help="train dir",
    )
    parser.add_argument("-m2", "--modelpath2", default="", help="train dir")
    parser.add_argument("-m3", "--modelpath3", default="", help="train dir")
    parser.add_argument("-m4", "--modelpath4", default="", help="train dir")
    parser.add_argument("-m5", "--modelpath5", default="", help="train dir")
    parser.add_argument("-v", "--validdir", default="", help="valid dir")
    parser.add_argument("-t", "--thr", default=32768, type=int, help="valid dir")
    parser.add_argument("-z", "--compare", default=0, type=int, help="sd")
    parser.add_argument("-d", "--df", default=False, type=bool, help="delflag")
    parser.add_argument("-c", "--c", default=True, type=bool, help="testflag")
    parser.add_argument("-mn", "--mn", default="", help="train dir")
    parser.add_argument("-f", "--flag", default="", help="specific version")
    parser.add_argument("-ver", "--tbd_ver", default="", help="specific version")
    parser.add_argument("-note", "--tbd_note", default="", help="specific version")
    parser.add_argument("-name", "--tbd_name", default="", help="specific version")

    args = parser.parse_args()
    return args


class Dataset(torchvision.datasets.ImageFolder):
    def __init__(
        self,
        root,
        transform=None,
        target_transform=None,
        loader=None,
        is_valid_file=None,
    ):
        super(Dataset, self).__init__(
            root,
            transform=transform,
            target_transform=target_transform,
            is_valid_file=is_valid_file,
        )

    def __getitem__(self, index):
        path, target = self.samples[index]
        sample = self.loader(path)
        # if "部分按压" not in path:
        if self.transform is not None:
            sample = self.transform(sample)
        if self.target_transform is not None:
            target = self.target_transform(target)

        return sample, target, path


def plot_confusion_matrix(
    cm, classes, normalize=False, title="Confusion matrix", cmap=plt.cm.Blues
):
    """
    This function prints and plots the confusion matrix.
    Normalization can be applied by setting `normalize=True`.
    Input
    - cm : 计算出的混淆矩阵的值
    - classes : 混淆矩阵中每一行每一列对应的列
    - normalize : True:显示百分比, False:显示个数
    """
    if normalize:
        cm = cm.astype("float") / cm.sum(axis=1)[:, np.newaxis]
        print("Normalized confusion matrix")
    else:
        print("Confusion matrix, without normalization")
    print(cm.shape)
    for i in range(cm.shape[0]):
        print("\n")
        for j in range(2):
            print("{:.5f}".format(cm[i][j]), end=" ")
    fig = plt.figure(figsize=(15, 10))
    ax = fig.add_subplot(111)
    ax.xaxis.set_ticks_position("top")
    tick_marks = np.arange(len(classes))

    # plt.tick_params(labelsize=13)
    plt.xticks(tick_marks, classes, rotation=0, fontsize=15)
    plt.yticks(tick_marks, classes, fontsize=15)

    plt.imshow(cm, interpolation="nearest", cmap=cmap)
    plt.title(title)
    plt.colorbar()

    # fmt = '.2f' if normalize else 'd'
    thresh = cm.max() / 2.0
    for i, j in itertools.product(range(cm.shape[0]), range(cm.shape[1])):
        if normalize == True:
            string = "{:.2%}".format(cm[i][j])
        else:
            string = "{:}".format(cm[i][j])
        plt.text(
            j,
            i,
            string,
            fontsize=17,
            horizontalalignment="center",
            color="white" if cm[i, j] > thresh else "black",
        )
    plt.tight_layout()
    plt.ylabel("True label")
    plt.xlabel("Predicted label")


def calc_idx_valid(idx_v, cnt_m):
    """
    docstring
    """
    if idx_v == 0:
        return 1
    else:
        return idx_v * (4 + 3 * cnt_m) + 1
    pass


def Resize_cv2(img, size):
    h, w = size
    img = np.array(img)
    img = cv2.resize(img, (w, h))
    img = Image.fromarray(img)
    # img = np.array(img)
    # print(img)
    return img


def test(args, idx, idm, idv, flag_test, thr, thr_hard, row_valid):
    # thr=50000
    # thr_hard=40000
    modelpath = args.listmodelpath[idm]
    validdir = args.listvaliddir[idv]
    time = time_ori + str(idx) + "_" + args.flag
    log_pth = modelpath.replace(".pth", "").split("/")[-1]
    log_test = validdir.split("/")[-1]
    time = time + log_pth + log_test
    device = torch.device("cuda")

    checkpoint = torch.load(modelpath)
    # checkpoint_h = torch.load(args.jilianmodelpath)
    jilianmodelpath = args.jilianmodelpath[idm]
    checkpoint_h = torch.load(jilianmodelpath)
    print(
        "\033[1;35mmodel path:{} jilianpath:{}\033[0m 模型{}, 测试集{}\033[0m".format(
            modelpath, args.jilianmodelpath, idm, idv
        )
    )

    print("\033[1;35mthreshold:{}\033[0m".format(thr))

    # G = MNV3_large2(2).cuda()

    # net=ASP0809().to(device)
    net = MNV3_large2_mixup().to(device)
    net_h = MNV3_large2_mixup().to(device)
    print("\033[1;35m thr:{} thr_hard:{}\033[0m  \n".format(thr, thr_hard))
    net.load_state_dict(checkpoint["net"])
    net_h.load_state_dict(checkpoint_h["net"])
    model_parameters = filter(lambda p: p.requires_grad, net.parameters())
    params = sum([np.prod(p.size()) for p in model_parameters])
    print(params)
    # ddd
    softmax = nn.Softmax(dim=1)

    # test_dir=r'/home/panq/dataset/spoof/6250/confirm_score_sz2'
    # test_dir=r'/hdd/file-input/wangb/classify/image/6191spoof/valid'
    # test_dir=r'/hdd/file-input/wangb/classify/image/6191spoof/confirm_score'
    test_dir = validdir
    print("\033[1;35mtest dir:{}\033[0m".format(test_dir))
    # test_dir=r'D:\doodle_lin\vivo\0618\test134'
    # test_dir=r'D:/doodle_lin/vivo/0618/vivo原图134/netTest'
    confusion_out_path = "./confusion_test/confusion_test%s" % time
    if not os.path.isdir(confusion_out_path):
        os.mkdir(confusion_out_path)
    # print(test_dir)
    test_transform = transforms.Compose(
        [  # transforms.Resize(config.input_size),
            # transforms.CenterCrop([153,153]),
            transforms.Grayscale(),
            # transforms.CenterCrop(config.input_size),
            #  transforms.ColorJitter(brightness=(0.8,1.2),contrast=(0.8,1.2)),
            transforms.ToTensor(),
            #  transforms.Normalize(mean=0.5, std=0.5)
            # transforms.Normalize()
        ]
    )
    # test_set = torchvision.datasets.ImageFolder(test_dir, test_transform)
    test_set = Dataset(test_dir, test_transform)
    if flag_test:
        print((test_set))
    dict = test_set.class_to_idx
    if flag_test:
        print(dict)
    classnum = len(test_set.classes)
    appear_times = Variable(torch.zeros(classnum, 1))
    for label in test_set.targets:
        appear_times[label] += 1
    confusionmap = Variable(torch.zeros(classnum, classnum))
    confusionmap_h = Variable(torch.zeros(classnum, classnum))
    confusionmap_jl = Variable(torch.zeros(classnum, classnum))
    val_loader = torch.utils.data.DataLoader(
        test_set, batch_size=2048, shuffle=False, num_workers=16
    )  # , pin_memory=True
    net.eval()
    net_h.eval()
    num = 0
    flag_1 = 0
    flag_0 = 0
    log = []
    dist_p_sum, dist_p_avg, dist_n_sum, dist_n_avg = 0, 0, 0, 0
    to_pil_image = transforms.ToPILImage()
    img_num = 0
    finger_num = 0
    miss_num = 0
    omit_num = 0
    idx_img_wuka = 2
    idx_img_louka = 2
    with torch.no_grad():
        with tqdm(total=len(val_loader), position=0, ncols=80) as pbar:
            (
                total,
                test_loss,
                test_correct,
                test_loss_ret,
                correct,
                correct_hard,
                correct_jilian,
                size,
                correct0,
                correct0_hard,
                correct0_jilian,
                size0,
                correct1,
                correct1_hard,
                correct1_jilian,
                size1,
            ) = (0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)

            for batch_num, (data, target, path) in enumerate(val_loader):
                data, target = data.to(device), target.to(device)

                target2 = [0.0 if (i < split_pos[0]) else 1.0 for i in target]
                target2 = torch.tensor(target).long().to(device)
                output = net(data, None, None, mode="test")
                output_h = net_h(data, None, None, mode="test")
                pred = torch.max(output, 1)
                pred_hard = torch.max(output_h, 1)
                for i in range(len(target)):
                    img_num += 1
                    img_type = path[i].split("/")
                    # if img_type[3] == '0fp':
                    # if '0fp' in path[i]:
                    if target[i] == 0:
                        finger_num += 1
                    # print(path[i])
                    log.append(
                        "\n%d: output[%d][0]:%f  output[%d][1]:%f path: %s \n"
                        % (i, i, output[i][0], i, output[i][1], path[i])
                    )

                    # if output[i][0] > output[i][1]:
                    #     c=output[i][0]
                    # else:
                    #     c=output[i][1]
                    # finger_score = exp(output[i][0] -c)/(exp(output[i][0]-c)+exp(output[i][1]-c))
                    p = softmax(output)
                    p_h = softmax(output_h)
                    finger_score = p[i][0]
                    finger_score_h = p_h[i][0]
                    spoof_score = p[i][1]

                    if thr_mode == 1:
                        pred.indices[i] = finger_score * 65536 < thr
                        pred_hard.indices[i] = finger_score_h * 65536 < thr
                    confusionmap[target[i]][pred.indices[i]] += 1
                    confusionmap_h[target[i]][pred_hard.indices[i]] += 1
                    if pred.indices[i] == 0 and pred_hard.indices[i] == 0:
                        confusionmap_jl[target[i]][0] += 1
                    else:
                        confusionmap_jl[target[i]][1] += 1
                        # prediction.indices[i]==1:#target2[i] != prediction.indices[i]
                    # print("path: "+path[i]+" score: "+str(finger_score*65536))
                    if 1:
                        if imgsort:
                            target_path = "./confusion_test/confusion_test%s/%d" % (
                                time,
                                target[i],
                            )
                            # if not os.path.isdir(target_path):
                            # os.mkdir(target_path)

                            img_type = path[i].split("/")
                            pathout = path[i].split("train/")[-1].replace("/", "_")
                            pathout = pathout[30:]
                            # print(pathout)
                            if target[i] == 0:
                                out_path = target_path
                                miss_num += 1
                            else:
                                out_path = target_path
                                omit_num += 1
                            if not os.path.isdir(out_path):
                                os.makedirs(out_path)
                            # copyfile(path[i], out_path + "/%03d.bmp"%num)
                            # copyfile(path[i], out_path + "/%d_%s"%(finger_score*65536,img_type[-1]))
                            # copyfile(path[i], out_path + "/%d_%s"%(finger_score*65536,pathout))
                            copyfile(
                                path[i],
                                out_path
                                + "/%d_%d_%s"
                                % (finger_score * 65536, finger_score_h, pathout)
                                #   path[i], out_path + "/%s_%d.bmp" %(pathout, finger_score * 65536)
                            )
                            img = pyxlImage(path[i])
                            if (
                                (finger_score * 65536) < 5000
                                # and target2[i] != pred.indices[i]
                                and idx_img_wuka <= 100
                            ):
                                idx_img_wuka += 1
                                _from = AnchorMarker(
                                    2 * ((idx + 1) * (idm + 1) + alphabet.index("A"))
                                    - 1,
                                    50000,
                                    str(idx_img_wuka - 1),
                                    50000,
                                )  # 列，大小，行，大小
                                _to = AnchorMarker(
                                    2 * ((idx + 1) * (idm + 1) + alphabet.index("A")),
                                    50000,
                                    str(idx_img_wuka),
                                    50000,
                                )
                                img.anchor = TwoCellAnchor("twoCell", _from, _to)
                                ws_wuka.add_image(img)
                                ws_wuka.cell(
                                    row=idx_img_wuka,
                                    column=2
                                    * ((idx + 1) * (idm + 1) + alphabet.index("A")),
                                    value=int(finger_score * 65536),
                                )
                            if (
                                (finger_score * 65536) > 30000
                                # and target2[i] != pred.indices[i]
                                and idx_img_louka <= 100
                            ):
                                idx_img_louka += 1
                                afrom = AnchorMarker(
                                    2 * ((idx + 1) * (idm + 1) + alphabet.index("A"))
                                    - 1,
                                    50000,
                                    str(idx_img_louka - 1),
                                    50000,
                                )  # 列，大小，行，大小
                                to = AnchorMarker(
                                    2 * ((idx + 1) * (idm + 1) + alphabet.index("A")),
                                    50000,
                                    str(idx_img_louka),
                                    50000,
                                )
                                img.anchor = TwoCellAnchor("twoCell", afrom, to)
                                ws_louka.add_image(img)
                                ws_louka.cell(
                                    row=idx_img_louka,
                                    column=2
                                    * ((idx + 1) * (idm + 1) + alphabet.index("A"))
                                    - 1,
                                    value=int(finger_score * 65536),
                                )

                            # print(finger_score*65536)
                        # if target[i] == 0:
                        # os.remove(path[i])
                        # copyfile(path[i], out_path + "/%s"%(img_type[-1]))
                        # move(path[i], out_path + "/%s"%(img_type[-1]))
                        # os.remove(path[i])
                        if logprint:
                            log.append(
                                "\ntarget:%d  num:%d path: %s \n"
                                % (target[i], num, path[i])
                            )
                        num += 1

                pred = pred[1]
                pred_hard = pred_hard[1]

                pred_jilian = [
                    0 if (i == j and i == 0) else 1 for (i, j) in zip(pred, pred_hard)
                ]
                pred_jilian = torch.tensor(pred_jilian).to(device)

                k = target.data.size()[0]
                correct += pred.eq(target.data).cpu().sum()
                correct_hard += pred_hard.eq(target.data).cpu().sum()
                correct_jilian += pred_jilian.eq(target.data).cpu().sum()
                size += k
                correct0 += (
                    torch.tensor(
                        [
                            1 if (i == j and i == 0) else 0
                            for (i, j) in zip(pred, target)
                        ]
                    )
                    .cpu()
                    .sum()
                )
                correct0_hard += (
                    torch.tensor(
                        [
                            1 if (i == j and i == 0) else 0
                            for (i, j) in zip(pred_hard, target)
                        ]
                    )
                    .cpu()
                    .sum()
                )
                correct0_jilian += (
                    torch.tensor(
                        [
                            1 if (i == j and i == 0) else 0
                            for (i, j) in zip(pred_jilian, target)
                        ]
                    )
                    .cpu()
                    .sum()
                )
                size0 += -(target - 1).cpu().sum()
                correct1 += (
                    torch.tensor(
                        [
                            1 if (i == j and i == 1) else 0
                            for (i, j) in zip(pred, target)
                        ]
                    )
                    .cpu()
                    .sum()
                )
                correct1_hard += (
                    torch.tensor(
                        [
                            1 if (i == j and i == 1) else 0
                            for (i, j) in zip(pred_hard, target)
                        ]
                    )
                    .cpu()
                    .sum()
                )
                correct1_jilian += (
                    torch.tensor(
                        [
                            1 if (i == j and i == 1) else 0
                            for (i, j) in zip(pred_jilian, target)
                        ]
                    )
                    .cpu()
                    .sum()
                )
                size1 += target.cpu().sum()

                pbar.update(1)

    notfinger = img_num - finger_num
    cm = confusionmap.numpy()
    cm_h = confusionmap_h.numpy()
    cm_jl = confusionmap_jl.numpy()
    if args.c:
        label = [0 if (i < split_pos[0]) else 1 for i in range(cm.shape[0])]
    else:
        label = [0 if (i < split_pos[0]) else 1 for i in range(cm.shape[0])]
    label = [0 if (i < split_pos[0]) else 1 for i in range(cm.shape[0])]
    cm_cor = torch.max(torch.Tensor(cm), 1)
    cm_all = cm.sum(axis=1)[:, np.newaxis]
    cm_cor_h = torch.max(torch.Tensor(cm_h), 1)
    cm_all_h = cm_h.sum(axis=1)[:, np.newaxis]
    cm_cor_jl = torch.max(torch.Tensor(cm_jl), 1)
    cm_all_jl = cm_jl.sum(axis=1)[:, np.newaxis]
    cm = cm_cor[0].data.numpy().astype("float") / cm.sum(axis=1)
    cm_h = cm_cor_h[0].data.numpy().astype("float") / cm_h.sum(axis=1)
    cm_jl = cm_cor_jl[0].data.numpy().astype("float") / cm_jl.sum(axis=1)
    print("|图库|比例|准确率|")
    print("|---|---|---|")
    global flag_name1

    idx_col = 4
    for i in range(cm.shape[0]):
        if int(cm_all[i][0]):
            print(
                "|{:<20}".format(list(dict.keys())[list(dict.values()).index(i)]),
                end="",
            )
            string_temp = "\nNORM |{}/{}|\033[1;33m {:.2f}%\033[0m|class {}".format(
                int(cm_cor[0][i]), int(cm_all[i][0]), cm[i] * 100, cm_cor[1][i]
            )
            string_temp_h = "HARD |{}/{}|\033[1;33m {:.2f}%\033[0m|class {}".format(
                int(cm_cor_h[0][i]), int(cm_all_h[i][0]), cm_h[i] * 100, cm_cor_h[1][i]
            )
            string_temp_jl = "JILI |{}/{}|\033[1;33m {:.2f}%\033[0m|class {}".format(
                int(cm_cor_jl[0][i]),
                int(cm_all_jl[i][0]),
                cm_jl[i] * 100,
                cm_cor_jl[1][i],
            )
            print("{:<40}".format(string_temp))
            print("{:<40}".format(string_temp_h))
            print("{:<40}".format(string_temp_jl))
            if flag_name1 == 0:
                # sheet.write(1, idx_col, list(dict.keys())[list(dict.values()).index(i)])  # 图库
                t1 = ws1.cell(
                    row=row_valid,
                    column=idx_col,
                    value=list(dict.keys())[list(dict.values()).index(i)],
                )
                # sheet.write_number(2, idx_col, int(cm_all[i][0]))  # number
                t2 = ws1.cell(
                    row=row_valid + 1, column=idx_col, value=int(cm_all[i][0])
                )

                str_01 = "0fp" if not cm_cor[1][i] else "1notfp"
                t3 = ws1.cell(row=row_valid + 2, column=idx_col, value=str_01)
                t1.style = "20 % - Accent1"
                t2.style = "20 % - Accent1"
                t3.style = "20 % - Accent1"
                # sheet.write(3, idx_col, str_01)  # class

            t4 = ws1.cell(
                row=row_valid + 3 * idx_m + 3, column=idx_col, value=cm[i]
            )  # number
            t5 = ws1.cell(
                row=row_valid + 3 * idx_m + 3 + 1, column=idx_col, value=cm_h[i]
            )  # number
            t6 = ws1.cell(
                row=row_valid + 3 * idx_m + 3 + 2, column=idx_col, value=cm_jl[i]
            )  # number

            if ws1.cell(row=row_valid + 2, column=idx_col).value == "0fp":
                t4.style = "20 % - Accent3"
            else:
                t6.style = "20 % - Accent3"
            idx_col += 1

    # for i in range(cm.shape[0]):

    if finger_num == 0:
        finger_num = 1
    if notfinger == 0:
        notfinger = 1
    # print("accuracy:%f  miss_percent:%f  omit_percent:%f "%((img_num-miss_num-omit_num)/img_num*100,miss_num/finger_num*100, (omit_num)/(notfinger)*100))
    log.append(
        "\nimg_num:%d  finger_num:%d  miss_num:%d omit_num: %d \n"
        % (img_num, finger_num, miss_num, omit_num)
    )
    log.append(
        "\naccuracy:%f  miss_percent:%f  omit_percent:%f \n"
        % (
            (img_num - miss_num - omit_num) / img_num * 100,
            miss_num / finger_num * 100,
            (omit_num) / (notfinger) * 100,
        )
    )
    if logprint:
        f = open(confusion_out_path + "/log.txt", "a")
        f.writelines(log)
        f.close()
    return


if __name__ == "__main__":
    softmax = nn.Softmax(dim=1)
    # ti.sleep(40000)
    # global args
    args = parse_args()
    softmax = nn.Softmax(dim=1)
    time_ori = "{}_{}".format(
        datetime.datetime.now().strftime("_%Y_%m_%d_%H_%M"), random_str(8)
    )
    # print(time)
    imgsort = 1
    logprint = 1
    thr_mode = 1
    solo_mode = 1
    # thr = args.thr
    split_pos = [1, 2, 3]
    delFlag = args.df
    teacher_flag = 0
    # netFlag=0
    copyEmptyDir = 0
    if delFlag:
        print("original picture has been moved!")
    device = torch.device("cuda")
    cnt = 0

    tbd_ver = "norm404079300 hard 404079300"
    tbd_note = "测试自动生成checklist"
    tbd_name = "潘琪"
    path_xlsx = args.xlsxpath + tbd_ver + time_ori + ".xlsx"

    for idx_v in range(len(args.listvaliddir)):
        validdir = args.listvaliddir[idx_v]
        path_validdir = Path(validdir)
        path_valid = (
            "/".join(path_validdir.parts[-2:]) if path_validdir.parts else path_validdir
        )
        print("path_valid: " + str(path_valid))
        # row_valid = calc_idx_valid(idx_v, len(args.listmodelpath))
        row_valid = idx_v * (4 + 3 * len(args.listmodelpath)) + 1

        ws1.cell(row=row_valid, column=1, value=str(path_valid))
        ws1["A" + str(row_valid)].style = "Check Cell"
        ws1.cell(row=row_valid, column=2, value="网络层")
        ws1.cell(row=row_valid, column=3, value="图库名称")
        ws1.cell(row=row_valid + 1, column=3, value="张数")
        ws1.cell(row=row_valid + 2, column=3, value="类型")

        list_valid = os.listdir(validdir)
        str_right = alphabet[alphabet.index("C") + len(list_valid)]

        temp_row = row_valid + 2
        # set_outer_border(
        #     ws1,
        #     "C"
        #     + str(row_valid)
        #     + ":"
        #     + str_right
        #     + str(row_valid + 3 * len(args.listmodelpath) + 3 + 2),
        #     0,
        # )

        ws1.merge_cells("A" + str(row_valid) + ":A" + str(temp_row))
        ws1.merge_cells("B" + str(row_valid) + ":B" + str(temp_row))
        flag_name1 = 0
        for idx_m in range(len(args.listmodelpath)):
            thr = 7000
            thr_hard = 30000

            modelpath = args.listmodelpath[idx_m]
            jilianmodelpath = args.jilianmodelpath[idx_m]
            path_modelpath = Path(modelpath)
            path_jilianmodelpath = Path(jilianmodelpath)
            path_model = (
                "/".join(path_modelpath.parts[-2:])
                if path_modelpath.parts
                else path_modelpath
            )
            path_jilianmodel = (
                "/".join(path_jilianmodelpath.parts[-2:])
                if path_jilianmodelpath.parts
                else path_jilianmodelpath
            )
            print("path_model: " + str(path_model) + str(path_jilianmodel))

            ws1.cell(
                row=row_valid + 3 * idx_m + 3,
                column=1,
                value=str(path_model) + str(path_jilianmodel),
            )
            ws1.cell(row=row_valid + 3 * idx_m + 3, column=2, value="norm")
            ws1.cell(row=row_valid + 3 * idx_m + 3, column=3, value=thr)
            ws1.cell(row=row_valid + 3 * idx_m + 3 + 1, column=2, value="hard")
            ws1.cell(row=row_valid + 3 * idx_m + 3 + 1, column=3, value=thr_hard)
            ws1.cell(row=row_valid + 3 * idx_m + 3 + 2, column=2, value="级联")
            ws1.merge_cells(
                "A"
                + str(row_valid + 3 * idx_m + 3)
                + ":A"
                + str(row_valid + 3 * idx_m + 3 + 2)
            )
            ws_wuka.cell(
                row=1, column=2 * (idx_m + 1) * (idx_v + 1), value=str(path_model)
            )
            ws_wuka.cell(
                row=2, column=2 * (idx_m + 1) * (idx_v + 1), value=str(path_valid)
            )
            ws_louka.cell(
                row=1, column=2 * (idx_m + 1) * (idx_v + 1), value=str(path_model)
            )
            ws_louka.cell(
                row=2, column=2 * (idx_m + 1) * (idx_v + 1), value=str(path_valid)
            )
            test(args, cnt, idx_m, idx_v, 0, thr, thr_hard, row_valid)
            set_outer_border(
                ws1,
                "C"
                + str(row_valid + 3 * idx_m + 3)
                + ":"
                + str_right
                + str(row_valid + 3 * idx_m + 3 + 2),
                1,
            )

        set_outer_border(ws1, "C" + str(row_valid) + ":" + str_right + str(temp_row), 0)

        set_inter_align(
            ws1,
            "A1"
            + ":"
            + str_right
            + str(row_valid + 3 * len(args.listmodelpath) + 3 + 2),
        )
        set_inter_percent(
            ws1,
            "D"
            + str(row_valid + 2)
            + ":"
            + str_right
            + str(row_valid + 3 * len(args.listmodelpath) + 3 + 2),
        )

    cnt_row = (
        (len(args.listmodelpath) + 1) * 3 * len(args.listvaliddir)
        + len(args.listvaliddir)
        - 1
    )

    # Set column width and row height for all cells

    ws1["C4"].font = Font(color="FF0000")
    ws1["C5"].font = Font(color="FF0000")

    set_auto_hw(ws_wuka, 14, 85)  # 180*200
    set_auto_hw(ws_louka, 14, 85)
    write_ws0(ws0)
    tbd_path_save_xlsx = (
        "/home/panq/tool/aosp/vendor/confusion_xlsx/算法网络checklist_"
        + time_str
        + tbd_ver
        + ".xlsx"
    )
    wb.save(tbd_path_save_xlsx)

    print("表格文件: " + path_xlsx)
